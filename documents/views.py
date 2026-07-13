""" from documents.services.ai_extractor import extract_information

from documents.services.framework_engine import evaluate_framework

from documents.services.scoring_engine import calculate_score

from documents.services.excel_generator import generate_excel

from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.shortcuts import render, redirect
import os
from documents.services.rule_engine import evaluate_rules
from .forms import DocumentUploadForm

from documents.models import (
    Document,
    DocumentChunk,
    ChatHistory
)

from documents.services.ocr_service import extract_pdf_text

from documents.utils.chunker import chunk_text
from documents.utils.embedding import create_embedding
from documents.utils.search import search_chunks
from documents.utils.chat import ask_gpt
from documents.models import Document
from documents.models import ChatHistory
from django.shortcuts import redirect
from documents.services.ocr_service import extract_text """

import os
import re
import requests
from urllib.parse import parse_qs, urlparse

from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect
from zipfile import BadZipFile, ZipFile
from openpyxl import load_workbook

from .forms import DocumentUploadForm

from documents.models import (
    Document,
    DocumentChunk,
    ChatHistory,
)

from documents.services.ai_extractor import extract_information
from documents.services.rule_engine import evaluate_rules
from documents.services.scoring_engine import calculate_score
from documents.services.excel_generator import generate_excel
from documents.services.ocr_service import extract_text

from documents.utils.chunker import chunk_text
from documents.utils.embedding import create_embedding
from documents.utils.search import search_chunks
from documents.utils.chat import ask_gpt

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import json


def process_saved_pdf(document):

    print("\n")
    print("=" * 70)
    print("ASSESSMENT STARTED")
    print("=" * 70)
    print("Framework    :", document.framework.code)
    print("Document     :", document.title)
    print("Document Type:", document.document_type)
    print("File Path    :", document.file.path)
    print("=" * 70)

    text = extract_text(
        document.file.path,
        document.document_type
    )

    document.extracted_text = text

    if text:

        print("\n")
        print("=" * 70)
        print("OCR COMPLETED")
        print("=" * 70)
        print("Characters Extracted :", len(text))
        print("=" * 70)

        extracted_data = extract_information(
            text,
            document.framework
        )

        print("\n")
        print("=" * 70)
        print("AI EXTRACTION")
        print("=" * 70)
        print(extracted_data)
        print("=" * 70)

        evaluation = evaluate_rules(
            document.framework,
            extracted_data
        )

        print("\n")
        print("=" * 70)
        print("RULE ENGINE RESULT")
        print("=" * 70)
        print(evaluation)
        print("=" * 70)

        category = evaluation.get(
            "category",
            ""
        )

        days = int(
            extracted_data.get(
                "duration_days",
                0
            )
        )

        print("\n")
        print("=" * 70)
        print("SCORING INPUT")
        print("=" * 70)
        print("Category :", category)
        print("Days     :", days)
        print("=" * 70)

        scoring = calculate_score(
            document.framework,
            category,
            days
        )

        print("\n")
        print("=" * 70)
        print("SCORING RESULT")
        print("=" * 70)
        print(scoring)
        print("=" * 70)

        excel_path = generate_excel(
            document=document,
            framework=document.framework,
            extracted_data=extracted_data,
            evaluation=evaluation,
            scoring=scoring
        )

        print("\n")
        print("=" * 70)
        print("EXCEL GENERATED")
        print("=" * 70)
        print("Excel Path :", excel_path)
        print("=" * 70)

        document.status = "Processed"

        chunks = chunk_text(text)

        for chunk in chunks:

            try:

                vector = create_embedding(chunk)

                DocumentChunk.objects.create(
                    document=document,
                    chunk_text=chunk,
                    embedding=vector,
                    page_number=1
                )

            except Exception as e:

                print("Embedding Error:", e)

        print("\n")
        print("=" * 70)
        print("EMBEDDING COMPLETED")
        print("=" * 70)
        print("Total Chunks :", len(chunks))
        print("=" * 70)

        print("\n")
        print("=" * 70)
        print("ASSESSMENT COMPLETED")
        print("=" * 70)
        print("Framework :", document.framework.code)
        print("Document  :", document.title)
        print("Excel     :", excel_path)
        print("=" * 70)

        document.save(
            update_fields=[
                "status",
                "extracted_text"
            ]
        )

    else:

        document.status = "OCR Required"

        print("\n")
        print("=" * 70)
        print("OCR FAILED")
        print("=" * 70)
        print("No text could be extracted from the document.")
        print("=" * 70)

        document.save(
            update_fields=[
                "status",
                "extracted_text"
            ]
        )


def download_pdf_from_url(url):

    link = str(url).strip()
    if not link:
        raise ValueError("No URL provided")

    parsed = urlparse(link)
    if "drive.google.com" in parsed.netloc:
        file_id = None
        match = re.search(r"/file/d/([^/]+)", link)
        if match:
            file_id = match.group(1)
        else:
            query = parse_qs(parsed.query)
            file_id = query.get("id", [None])[0]

        if file_id:
            link = f"https://drive.google.com/uc?export=download&id={file_id}"

    response = requests.get(link, stream=True, timeout=30, allow_redirects=True)
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "")
    if "pdf" not in content_type.lower() and not link.lower().endswith(".pdf"):
        disposition = response.headers.get("Content-Disposition", "")
        if ".pdf" not in disposition.lower():
            raise ValueError("URL does not return a PDF file")

    filename = None
    disposition = response.headers.get("Content-Disposition", "")
    if "filename=" in disposition:
        names = re.findall(r'filename="?([^";]+)"?', disposition)
        if names:
            filename = names[0]

    if not filename:
        filename = os.path.basename(parsed.path) or "downloaded.pdf"
        if not filename.lower().endswith(".pdf"):
            filename = f"{filename}.pdf"

    return response.content, filename


def process_excel_file(uploaded_by, framework, title, excel_file):

    excel_file.seek(0)

    try:
        workbook = load_workbook(excel_file, data_only=True)
    except Exception:
        return -1

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    headers = [str(cell).strip().lower() if cell else "" for cell in header_row]

    accepted_headers = ["pdflink", "pdf link", "pdf_link", "link", "url"]
    link_col = None
    for index, header in enumerate(headers):
        if header in accepted_headers:
            link_col = index
            break

    if link_col is None:
        return -1

    processed_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        raw_url = row[link_col]
        if not raw_url:
            continue

        try:
            pdf_data, pdf_name = download_pdf_from_url(raw_url)
        except Exception as e:
            print("PDF download failed:", raw_url, e)
            continue

        pdf_title = title if title else pdf_name
        if title:
            pdf_title = f"{title} - {pdf_name}"

        pdf_file = ContentFile(pdf_data, name=pdf_name)

        extracted_document = Document(
            uploaded_by=uploaded_by,
            framework=framework,
            title=pdf_title,
            file=pdf_file,
            document_type="pdf"
        )
        extracted_document.save()
        process_saved_pdf(extracted_document)
        processed_count += 1

    return processed_count


def process_zip_file(uploaded_by, framework, title, zip_file):

    zip_file.seek(0)

    try:

        with ZipFile(zip_file) as archive:

            pdf_infos = [
                info for info in archive.infolist()
                if not info.is_dir() and info.filename.lower().endswith(".pdf")
            ]

            for info in pdf_infos:

                with archive.open(info) as zipped_file:

                    file_data = zipped_file.read()

                pdf_name = os.path.basename(info.filename) or "extracted.pdf"
                pdf_title = title if title else pdf_name
                if title:
                    pdf_title = f"{title} - {pdf_name}"

                extracted_file = ContentFile(
                    file_data,
                    name=pdf_name
                )

                extracted_document = Document(
                    uploaded_by=uploaded_by,
                    framework=framework,
                    title=pdf_title,
                    file=extracted_file,
                    document_type="pdf"
                )

                extracted_document.save()
                process_saved_pdf(extracted_document)

            return len(pdf_infos)

    except BadZipFile:

        return -1


def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def delete_documents(request):

    if request.method == "POST":

        ids = request.POST.getlist(
            "documents"
        )

        Document.objects.filter(
            id__in=ids
        ).delete()

    return redirect(
        "document_library"
    )
@login_required
def document_library(request):

    documents = Document.objects.all().order_by(
        "-uploaded_at"
    )

    return render(
        request,
        "documents/library.html",
        {
            "documents": documents
        }
    )

@login_required
def chat_history(request):

    chats = ChatHistory.objects.all().order_by(
        "-created_at"
    )

    return render(
        request,
        "documents/history.html",
        {
            "chats": chats
        }
    )

# =========================
# DOCUMENT UPLOAD
# =========================

@login_required
def upload_document(request):

    documents = Document.objects.order_by("-uploaded_at")
    if request.method == "POST":

        form = DocumentUploadForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            framework = form.cleaned_data["framework"]
            title = form.cleaned_data.get("title", "")
            uploaded_files = form.cleaned_data.get("files", [])

            for uploaded_file in uploaded_files:

                document = Document(
                    uploaded_by=request.user,
                    framework=framework,
                    title=title if title else uploaded_file.name,
                    file=uploaded_file
                )

                extension = os.path.splitext(
                    uploaded_file.name
                )[1].lower()

                if extension == ".pdf":
                    document.document_type = "pdf"

                elif extension in [
                    ".png",
                    ".jpg",
                    ".jpeg",
                    ".bmp",
                    ".tif",
                    ".tiff"
                ]:
                    document.document_type = "image"

                elif extension == ".docx":
                    document.document_type = "docx"

                elif extension == ".xlsx":
                    document.document_type = "xlsx"

                elif extension == ".pptx":
                    document.document_type = "pptx"

                elif extension == ".csv":
                    document.document_type = "csv"

                elif extension == ".txt":
                    document.document_type = "txt"

                elif extension == ".zip":
                    document.document_type = "zip"

                else:
                    document.document_type = "unknown"

                document.save()

                print("\n")
                print("=" * 70)
                print("ASSESSMENT STARTED")
                print("=" * 70)
                print("Framework    :", document.framework.code)
                print("Document     :", document.title)
                print("Document Type:", document.document_type)
                print("File Path    :", document.file.path)
                print("=" * 70)

                if document.document_type == "zip":
                    document.status = "Uploaded"
                    document.save(update_fields=["status"])

                    pdf_count = process_zip_file(
                        request.user,
                        framework,
                        title,
                        uploaded_file
                    )

                    if pdf_count == 0:
                        document.status = "No PDF Found"
                        document.save(update_fields=["status"])

                    elif pdf_count < 0:
                        document.status = "Invalid ZIP"
                        document.save(update_fields=["status"])

                    continue

                if document.document_type == "xlsx":
                    document.status = "Uploaded"
                    document.save(update_fields=["status"])

                    pdf_count = process_excel_file(
                        request.user,
                        framework,
                        title,
                        uploaded_file
                    )

                    if pdf_count == 0:
                        document.status = "No PDF Links Found"
                        document.save(update_fields=["status"])

                    elif pdf_count < 0:
                        document.status = "Invalid Excel"
                        document.save(update_fields=["status"])

                    continue

                # --------------------------------------------------
                # OCR
                # --------------------------------------------------

                text = extract_text(
                    document.file.path,
                    document.document_type
                )

                document.extracted_text = text

                if text:

                    print("\n")
                    print("=" * 70)
                    print("OCR COMPLETED")
                    print("=" * 70)
                    print("Characters Extracted :", len(text))
                    print("=" * 70)

                    # --------------------------------------------------
                    # AI Extraction
                    # --------------------------------------------------

                    extracted_data = extract_information(
                        text,
                        document.framework
                    )

                    print("\n")
                    print("=" * 70)
                    print("AI EXTRACTION")
                    print("=" * 70)
                    print(extracted_data)
                    print("=" * 70)

                    # --------------------------------------------------
                    # Rule Engine
                    # --------------------------------------------------

                    evaluation = evaluate_rules(
                        document.framework,
                        extracted_data
                    )

                    print("\n")
                    print("=" * 70)
                    print("RULE ENGINE RESULT")
                    print("=" * 70)
                    print(evaluation)
                    print("=" * 70)

                    # --------------------------------------------------
                    # Scoring Engine
                    # --------------------------------------------------

                    category = evaluation.get(
                        "category",
                        ""
                    )

                    days = int(
                        extracted_data.get(
                            "duration_days",
                            0
                        )
                    )

                    print("\n")
                    print("=" * 70)
                    print("SCORING INPUT")
                    print("=" * 70)
                    print("Category :", category)
                    print("Days     :", days)
                    print("=" * 70)

                    scoring = calculate_score(
                        document.framework,
                        category,
                        days
                    )

                    print("\n")
                    print("=" * 70)
                    print("SCORING RESULT")
                    print("=" * 70)
                    print(scoring)
                    print("=" * 70)

                    # --------------------------------------------------
                    # Excel Generation
                    # --------------------------------------------------

                    excel_path = generate_excel(
                        document=document,
                        framework=document.framework,
                        extracted_data=extracted_data,
                        evaluation=evaluation,
                        scoring=scoring
                    )

                    print("\n")
                    print("=" * 70)
                    print("EXCEL GENERATED")
                    print("=" * 70)
                    print("Excel Path :", excel_path)
                    print("=" * 70)

                    # --------------------------------------------------
                    # Embedding Generation
                    # --------------------------------------------------

                    document.status = "Processed"

                    chunks = chunk_text(text)

                    for chunk in chunks:

                        try:

                            vector = create_embedding(chunk)

                            DocumentChunk.objects.create(
                                document=document,
                                chunk_text=chunk,
                                embedding=vector,
                                page_number=1
                            )

                        except Exception as e:

                            print("Embedding Error:", e)

                    print("\n")
                    print("=" * 70)
                    print("EMBEDDING COMPLETED")
                    print("=" * 70)
                    print("Total Chunks :", len(chunks))
                    print("=" * 70)

                    print("\n")
                    print("=" * 70)
                    print("ASSESSMENT COMPLETED")
                    print("=" * 70)
                    print("Framework :", document.framework.code)
                    print("Document  :", document.title)
                    print("Excel     :", excel_path)
                    print("=" * 70)

                    document.save(
                        update_fields=[
                            "status",
                            "extracted_text"
                        ]
                    )

                else:

                    document.status = "OCR Required"

                    print("\n")
                    print("=" * 70)
                    print("OCR FAILED")
                    print("=" * 70)
                    print("No text could be extracted from the document.")
                    print("=" * 70)

                    document.save(
                        update_fields=[
                            "status",
                            "extracted_text"
                        ]
                    )

        else:

            print(form.errors)

            return render(
                request,
                "documents/upload.html",
                {
                    "form": form,
                    "documents": documents
                }
            )

    else:

        form = DocumentUploadForm()

    return render(
        request,
        "documents/upload.html",
        {
            "form": form,
            "documents": documents
        }
    )

    return render(
        request,
        "documents/upload.html",
        {
            "form": form,
            "documents": documents
        }
    )
# =========================
# CHAT WITH DOCUMENT
# =========================


#@login_required
@require_POST
@csrf_exempt
def chat_api(request):

    try:
        data = json.loads(request.body)

        question = data.get("question", "")
        document_id = data.get("document_id")
        question = data.get("question", "").strip()

        greetings = [
            "hi",
            "hello",
            "hey",
            "good morning",
            "good evening",
            "good afternoon"
        ]

        if question.lower() in greetings:
            return JsonResponse({
                "success": True,
                "answer": "Hello! 👋 How can I help you today? You can ask me anything about your uploaded documents."
            })
        query_embedding = create_embedding(question)

        results = search_chunks(
            query_embedding,
            document_id=document_id,
            limit=10
        )

        context = ""

        for chunk in results:
            context += f"""
Document:
{chunk.document.title}

Content:
{chunk.chunk_text}

"""

        answer = ask_gpt(question, context)

        ChatHistory.objects.create(
            question=question,
            answer=answer
        )

        return JsonResponse({
            "success": True,
            "answer": answer
        })

    except Exception as e:

        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)
        
        
@login_required
def ask_document(request):

    answer = ""

    if request.method == "POST":

        question = request.POST.get(
            "question"
        )

        document_id = request.POST.get(
            "document_id"
        )

        query_embedding = create_embedding(
            question
        )

        results = search_chunks(
            query_embedding,
            document_id=document_id,
            limit=10
        )

        context = ""

        for chunk in results:

            context += f"""

Document:
{chunk.document.title}

Content:
{chunk.chunk_text}

"""

        answer = ask_gpt(
            question,
            context
        )

        ChatHistory.objects.create(
            question=question,
            answer=answer
        )

    return render(
        request,
        "documents/chat.html",
        {
            "answer": answer,
            "documents": Document.objects.all(),
            "history": ChatHistory.objects.order_by(
                "-created_at"
            )[:20]
        }
    )