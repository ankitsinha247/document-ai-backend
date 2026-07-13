import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from django.conf import settings


def generate_excel(
    document,
    framework,
    extracted_data,
    evaluation,
    scoring
):

    wb = Workbook()

    ws = wb.active

    ws.title = framework.code

    # -----------------------------------
    # Header Row
    # -----------------------------------

    headers = [

        "faculty_name",

        "programme_name",

        "host_institution",

        "sponsoring_agency",

        "date",

        "duration",

        "days",

        "category",

        "weightage_per_day",

        "score",

        "eligible",

        "remarks"

    ]

    ws.append(headers)

    # Style headers
    blue_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = blue_fill

    # -----------------------------------
    # Date
    # -----------------------------------

    start = extracted_data.get("start_date", "")

    end = extracted_data.get("end_date", "")

    if start and end:

        date = f"{start} to {end}"

    else:

        date = extracted_data.get("date", "")

    # -----------------------------------
    # Days
    # -----------------------------------

    days = scoring.get(
        "days",
        extracted_data.get(
            "duration_days",
            0
        )
    )

    duration = f"{days} days"

    # -----------------------------------
    # Category
    # -----------------------------------

    category = scoring.get(
        "category",
        ""
    )

    category_number = (
        str(category)
        .replace("Category", "")
        .strip()
    )

    # -----------------------------------
    # Score
    # -----------------------------------

    weightage = scoring.get(
        "weightage_per_day",
        0
    )

    score = scoring.get(
        "total_score",
        0
    )

    # -----------------------------------
    # Remarks
    # -----------------------------------

    remarks = scoring.get(
        "remarks",
        ""
    )

    # -----------------------------------
    # Data Row
    # -----------------------------------

    ws.append([

        extracted_data.get(
            "faculty_name",
            ""
        ),

        extracted_data.get(
            "programme_name",
            ""
        ),

        extracted_data.get(
            "host_institution",
            ""
        ),

        extracted_data.get(
            "sponsoring_agency",
            ""
        ),

        date,

        duration,

        days,

        category_number,

        weightage,

        score,

        scoring.get(
            "eligible",
            False
        ),

        remarks

    ])

    # Auto-size columns
    for column_cells in ws.columns:

        length = max(
            len(str(cell.value or ""))
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 4, 50)

    # -----------------------------------
    # Save File
    # -----------------------------------

    folder = os.path.join(
        settings.MEDIA_ROOT,
        "reports"
    )

    os.makedirs(
        folder,
        exist_ok=True
    )

    filename = (
        f"{document.id}_"
        f"{framework.code}_"
        "assessment.xlsx"
    )

    filepath = os.path.join(
        folder,
        filename
    )

    wb.save(filepath)

    return filepath