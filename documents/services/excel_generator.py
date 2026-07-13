import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from django.conf import settings

from documents.models import OutputColumn


def generate_excel(
    document,
    framework,
    extracted_data,
    evaluation,
    scoring
):

    # --------------------------------------------------
    # Report Folder
    # --------------------------------------------------

    folder = os.path.join(
        settings.MEDIA_ROOT,
        "reports"
    )

    os.makedirs(
        folder,
        exist_ok=True
    )

    filename = f"{framework.code}_assessment.xlsx"

    filepath = os.path.join(
        folder,
        filename
    )

    # --------------------------------------------------
    # Read Output Columns
    # --------------------------------------------------

    columns = OutputColumn.objects.filter(
        framework=framework,
        is_active=True
    ).order_by(
        "column_order"
    )

    # --------------------------------------------------
    # Create Workbook / Open Existing
    # --------------------------------------------------

    if os.path.exists(filepath):

        wb = load_workbook(filepath)

        ws = wb.active

    else:

        wb = Workbook()

        ws = wb.active

        ws.title = framework.code

        headers = []

        for column in columns:

            headers.append(
                column.column_title
            )

        ws.append(headers)

        blue_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78"
        )

        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = blue_fill

    # --------------------------------------------------
    # Build Dynamic Row
    # --------------------------------------------------

    row = []

    for column in columns:

        value = ""

        # ------------------------------
        # Extracted Data
        # ------------------------------

        if column.source == "extracted":

            if column.field_name == "date":

                start = extracted_data.get(
                    "start_date",
                    ""
                )

                end = extracted_data.get(
                    "end_date",
                    ""
                )

                if start and end:

                    value = f"{start} to {end}"

                else:

                    value = extracted_data.get(
                        "date",
                        ""
                    )

            elif column.field_name == "duration":

                days = int(
                    extracted_data.get(
                        "duration_days",
                        0
                    )
                )

                value = f"{days} days"

            else:

                value = extracted_data.get(
                    column.field_name,
                    ""
                )

        # ------------------------------
        # Rule Engine
        # ------------------------------

        elif column.source == "evaluation":

            value = evaluation.get(
                column.field_name,
                ""
            )

            if column.field_name == "category":

                value = (
                    str(value)
                    .replace("Category", "")
                    .strip()
                )

        # ------------------------------
        # Scoring Engine
        # ------------------------------

        elif column.source == "scoring":

            value = scoring.get(
                column.field_name,
                ""
            )

        # ------------------------------
        # Document Model
        # ------------------------------

        elif column.source == "document":

            value = getattr(
                document,
                column.field_name,
                ""
            )

        row.append(value)

    ws.append(row)

    # --------------------------------------------------
    # Auto Width
    # --------------------------------------------------

    for column_cells in ws.columns:

        max_length = max(

            len(str(cell.value)) if cell.value else 0

            for cell in column_cells

        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(
            max_length + 3,
            60
        )

    # --------------------------------------------------
    # Save Workbook
    # --------------------------------------------------

    wb.save(filepath)

    return filepath